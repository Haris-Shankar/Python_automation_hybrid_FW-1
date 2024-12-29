import openpyxl


def get_row_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    row = sheet.max_row
    return row


def get_column_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    column = sheet.max_column
    return column


def read_data(file, sheetname, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    data = sheet.cell(row=rownum, column=colnum).value
    return data


def write_data(file, sheetname, rownum, colnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(rownum, colnum).value = data
    workbook.save(file)
